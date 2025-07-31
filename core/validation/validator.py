"""
Módulo de validación para Declaraciones Juradas.
Valida DataFrames contra las reglas definidas en la metadata de Access.
"""

import pandas as pd
import re
from typing import Dict, List, Any, Optional, Tuple
from datetime import datetime
import math
from ..access_schema import AccessSchema


class DJValidator:
    """Validador de Declaraciones Juradas."""
    
    def __init__(self, metadata: Dict[str, Any], access_schema: AccessSchema = None):
        """
        Inicializa el validador con metadata de la DJ.
        
        Args:
            metadata: Metadata completa de la DJ obtenida de Access
            access_schema: Instancia de AccessSchema para consultas adicionales
        """
        self.metadata = metadata
        self.access_schema = access_schema or AccessSchema()
        self.dj_codigo = metadata['declaracion']['dj_codigo']
        self.campos = metadata['campos']
        self.validaciones = metadata['validaciones']
        
        # Cache para tablas lookup
        self._lookup_cache = {}
    
    def validar_dataframe(self, df: pd.DataFrame) -> Dict[str, Any]:
        """
        Valida un DataFrame completo contra todas las reglas.
        
        Args:
            df: DataFrame a validar
            
        Returns:
            Diccionario estructurado con resultados de validación
        """
        resultado = {
            "valido": True,
            "errores": [],
            "resumen": {
                "total_filas": len(df),
                "errores_totales": 0,
                "columnas_con_error": []
            },
            "validaciones_omitidas": []
        }
        
        # 1. Verificar columnas obligatorias
        columnas_faltantes = self._verificar_columnas_obligatorias(df)
        if columnas_faltantes:
            resultado["valido"] = False
            for columna in columnas_faltantes:
                resultado["errores"].append({
                    "fila": None,
                    "columna": columna,
                    "codigo": "OBLIGATORIO",
                    "error": f"Columna obligatoria '{columna}' no está presente en el DataFrame"
                })
            # Si faltan columnas obligatorias, no continuar con otras validaciones
            resultado["resumen"]["errores_totales"] = len(resultado["errores"])
            return resultado
        
        # 2. Ejecutar validaciones por campo y fila
        columnas_con_error = set()
        
        for codigo_campo, validaciones_campo in self.validaciones.items():
            if codigo_campo not in df.columns:
                # Registrar validaciones omitidas
                for validacion in validaciones_campo:
                    resultado["validaciones_omitidas"].append({
                        "columna": codigo_campo,
                        "codigo": validacion["codigo_validacion"],
                        "motivo": "Columna no presente en el DataFrame"
                    })
                continue
            
            # Validar cada fila para este campo
            for idx, valor in enumerate(df[codigo_campo]):
                errores_fila = self._validar_campo_valor(
                    codigo_campo, valor, idx, df
                )
                
                if errores_fila:
                    resultado["errores"].extend(errores_fila)
                    columnas_con_error.add(codigo_campo)
                    resultado["valido"] = False
        
        # 3. Actualizar resumen
        resultado["resumen"]["errores_totales"] = len(resultado["errores"])
        resultado["resumen"]["columnas_con_error"] = list(columnas_con_error)
        
        return resultado
    
    def _verificar_columnas_obligatorias(self, df: pd.DataFrame) -> List[str]:
        """Verifica que estén presentes todas las columnas obligatorias."""
        columnas_faltantes = []
        
        for codigo_campo, info_campo in self.campos.items():
            if info_campo['obligatorio'] and codigo_campo not in df.columns:
                columnas_faltantes.append(codigo_campo)
        
        return columnas_faltantes
    
    def _validar_campo_valor(self, codigo_campo: str, valor: Any, fila_idx: int, df: pd.DataFrame) -> List[Dict[str, Any]]:
        """
        Valida un valor específico contra todas las reglas del campo.
        
        Args:
            codigo_campo: Código del campo (C1, C2, etc.)
            valor: Valor a validar
            fila_idx: Índice de la fila (0-based)
            df: DataFrame completo (para validaciones contextuales)
            
        Returns:
            Lista de errores encontrados
        """
        errores = []
        
        if codigo_campo not in self.validaciones:
            return errores
        
        validaciones_campo = self.validaciones[codigo_campo]
        
        for validacion in validaciones_campo:
            try:
                error = self._ejecutar_validacion(
                    validacion, codigo_campo, valor, fila_idx, df
                )
                if error:
                    errores.append(error)
            except Exception as e:
                # Error en la ejecución de la validación
                errores.append({
                    "fila": fila_idx + 1,  # 1-based para el usuario
                    "columna": codigo_campo,
                    "codigo": validacion["codigo_validacion"],
                    "error": f"Error en validación: {str(e)}"
                })
        
        return errores
    
    def _ejecutar_validacion(self, validacion: Dict[str, Any], codigo_campo: str, 
                           valor: Any, fila_idx: int, df: pd.DataFrame) -> Optional[Dict[str, Any]]:
        """
        Ejecuta una validación específica.
        
        Args:
            validacion: Diccionario con info de la validación
            codigo_campo: Código del campo
            valor: Valor a validar
            fila_idx: Índice de la fila
            df: DataFrame completo
            
        Returns:
            Diccionario con error si la validación falla, None si es válida
        """
        tipo_validacion = validacion["tipo_validacion"]
        expresion_py = validacion["expresion_py"]
        codigo_validacion = validacion["codigo_validacion"]
        mensaje_error = validacion["mensaje_error"]
        
        # Preparar contexto para eval()
        contexto = self._preparar_contexto_validacion(valor, fila_idx, df, codigo_campo)
        
        try:
            # Ejecutar la expresión Python
            resultado = eval(expresion_py, {"__builtins__": {}}, contexto)
            
            # Si el resultado es False, hay un error de validación
            if not resultado:
                return {
                    "fila": fila_idx + 1,  # 1-based para el usuario
                    "columna": codigo_campo,
                    "codigo": codigo_validacion,
                    "error": mensaje_error
                }
        except Exception as e:
            # Error en la evaluación de la expresión
            return {
                "fila": fila_idx + 1,
                "columna": codigo_campo,
                "codigo": codigo_validacion,
                "error": f"Error evaluando validación: {str(e)}"
            }
        
        return None
    
    def _preparar_contexto_validacion(self, valor: Any, fila_idx: int, df: pd.DataFrame, codigo_campo: str) -> Dict[str, Any]:
        """
        Prepara el contexto de variables para eval().
        
        Args:
            valor: Valor del campo actual
            fila_idx: Índice de la fila
            df: DataFrame completo
            codigo_campo: Código del campo actual
            
        Returns:
            Diccionario con variables disponibles para eval()
        """
        # Fila actual como Serie
        fila_actual = df.iloc[fila_idx] if fila_idx < len(df) else pd.Series()
        
        contexto = {
            # Valor actual
            'valor': valor,
            'v': valor,  # Alias corto
            
            # Información de fila
            'fila': fila_actual,
            'fila_idx': fila_idx,
            'fila_num': fila_idx + 1,  # 1-based
            
            # DataFrame completo
            'df': df,
            
            # Funciones útiles
            'len': len,
            'str': str,
            'int': int,
            'float': float,
            'bool': bool,
            'abs': abs,
            'round': round,
            'min': min,
            'max': max,
            'sum': sum,
            'any': any,
            'all': all,
            'isinstance': isinstance,
            'math': math,
            're': re,
            'pd': pd,
            'datetime': datetime,
            
            # Funciones personalizadas
            'es_nulo': lambda x: pd.isna(x) or x is None or (isinstance(x, str) and x.strip() == ''),
            'es_numerico': lambda x: isinstance(x, (int, float)) and not pd.isna(x),
            'es_texto': lambda x: isinstance(x, str),
            'longitud': lambda x: len(str(x)) if not pd.isna(x) else 0,
            'contiene': lambda texto, patron: str(patron) in str(texto) if not pd.isna(texto) else False,
            'coincide_regex': lambda texto, patron: bool(re.match(patron, str(texto))) if not pd.isna(texto) else False,
            'entre': lambda x, min_val, max_val: min_val <= x <= max_val if isinstance(x, (int, float)) and not pd.isna(x) else False,
            'en_lista': lambda x, lista: x in lista,
            'lookup': lambda tabla, campo_buscar, valor_buscar, campo_retorno: self._buscar_lookup(tabla, campo_buscar, valor_buscar, campo_retorno),
        }
        
        # Agregar acceso directo a otros campos de la fila actual
        for col in df.columns:
            if col != codigo_campo:  # Evitar referencia circular
                contexto[col.lower()] = fila_actual.get(col) if col in fila_actual else None
        
        return contexto
    
    def _buscar_lookup(self, tabla: str, campo_buscar: str, valor_buscar: Any, campo_retorno: str) -> Any:
        """
        Busca un valor en una tabla de referencia.
        
        Args:
            tabla: Nombre de la tabla lookup
            campo_buscar: Campo por el cual buscar
            valor_buscar: Valor a buscar
            campo_retorno: Campo a retornar
            
        Returns:
            Valor encontrado o None si no existe
        """
        try:
            # Cache de tablas lookup
            if tabla not in self._lookup_cache:
                self._lookup_cache[tabla] = self.access_schema.get_tabla_lookup(tabla)
            
            tabla_df = self._lookup_cache[tabla]
            
            # Buscar el registro
            resultado = tabla_df[tabla_df[campo_buscar] == valor_buscar]
            
            if len(resultado) > 0:
                return resultado.iloc[0][campo_retorno]
            else:
                return None
                
        except Exception:
            return None
    
    def generar_reporte_errores(self, resultado_validacion: Dict[str, Any], 
                               output_path: str = None) -> str:
        """
        Genera un reporte Excel con los errores de validación.
        
        Args:
            resultado_validacion: Resultado de validar_dataframe()
            output_path: Ruta donde guardar el archivo Excel
            
        Returns:
            Ruta del archivo generado
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = f"errores_validacion_DJ{self.dj_codigo}_{timestamp}.xlsx"
        
        wb = Workbook()
        
        # Hoja de resumen
        ws_resumen = wb.active
        ws_resumen.title = "Resumen"
        
        # Encabezados de resumen
        ws_resumen['A1'] = "REPORTE DE ERRORES DE VALIDACIÓN"
        ws_resumen['A1'].font = Font(bold=True, size=14)
        
        ws_resumen['A3'] = "DJ Código:"
        ws_resumen['B3'] = self.dj_codigo
        ws_resumen['A4'] = "Fecha:"
        ws_resumen['B4'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws_resumen['A5'] = "Estado:"
        ws_resumen['B5'] = "VÁLIDO" if resultado_validacion["valido"] else "CON ERRORES"
        
        # Resumen numérico
        resumen = resultado_validacion["resumen"]
        ws_resumen['A7'] = "Total de filas:"
        ws_resumen['B7'] = resumen["total_filas"]
        ws_resumen['A8'] = "Errores encontrados:"
        ws_resumen['B8'] = resumen["errores_totales"]
        ws_resumen['A9'] = "Columnas con error:"
        ws_resumen['B9'] = len(resumen["columnas_con_error"])
        
        # Hoja de errores detallados
        if resultado_validacion["errores"]:
            ws_errores = wb.create_sheet("Errores Detallados")
            
            # Encabezados
            encabezados = ["Fila", "Columna", "Código Error", "Descripción"]
            for col, encabezado in enumerate(encabezados, 1):
                ws_errores.cell(row=1, column=col, value=encabezado)
                ws_errores.cell(row=1, column=col).font = Font(bold=True)
                ws_errores.cell(row=1, column=col).fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Datos de errores
            for row, error in enumerate(resultado_validacion["errores"], 2):
                ws_errores.cell(row=row, column=1, value=error["fila"])
                ws_errores.cell(row=row, column=2, value=error["columna"])
                ws_errores.cell(row=row, column=3, value=error["codigo"])
                ws_errores.cell(row=row, column=4, value=error["error"])
            
            # Ajustar anchos de columna
            for column in ws_errores.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_errores.column_dimensions[column_letter].width = adjusted_width
        
        # Hoja de validaciones omitidas
        if resultado_validacion["validaciones_omitidas"]:
            ws_omitidas = wb.create_sheet("Validaciones Omitidas")
            
            # Encabezados
            encabezados = ["Columna", "Código", "Motivo"]
            for col, encabezado in enumerate(encabezados, 1):
                ws_omitidas.cell(row=1, column=col, value=encabezado)
                ws_omitidas.cell(row=1, column=col).font = Font(bold=True)
            
            # Datos
            for row, omitida in enumerate(resultado_validacion["validaciones_omitidas"], 2):
                ws_omitidas.cell(row=row, column=1, value=omitida["columna"])
                ws_omitidas.cell(row=row, column=2, value=omitida["codigo"])
                ws_omitidas.cell(row=row, column=3, value=omitida["motivo"])
        
        # Guardar archivo
        wb.save(output_path)
        return output_path


def validar_dj(df: pd.DataFrame, dj_codigo: str, db_path: str = None) -> Dict[str, Any]:
    """
    Función de conveniencia para validar un DataFrame.
    
    Args:
        df: DataFrame a validar
        dj_codigo: Código de la declaración jurada
        db_path: Ruta opcional al archivo Access
        
    Returns:
        Resultado de la validación
    """
    from ..access_schema import obtener_metadata
    
    # Obtener metadata
    metadata = obtener_metadata(dj_codigo, db_path)
    
    # Crear validador y ejecutar
    validator = DJValidator(metadata)
    return validator.validar_dataframe(df)


if __name__ == "__main__":
    # Ejemplo de uso
    import pandas as pd
    
    # Datos de ejemplo
    df_ejemplo = pd.DataFrame({
        'C1': ['12345678-9', '98987654-3', '11111111-1'],
        'C2': [1000, 2000, -100],  # Valor negativo para probar validación
        'C3': ['EMPRESA A', 'EMPRESA B', 'EMPRESA C']
    })
    
    try:
        resultado = validar_dj(df_ejemplo, "1922")
        
        if resultado["valido"]:
            print("✓ DataFrame válido")
        else:
            print("✗ DataFrame con errores:")
            for error in resultado["errores"]:
                print(f"  - Fila {error['fila']}, Campo {error['columna']}: {error['error']}")
                
    except Exception as e:
        print(f"Error en validación: {e}")
