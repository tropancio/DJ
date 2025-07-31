"""
Módulo para generar plantillas Excel de carga de datos.
Genera archivos Excel con estructura, comentarios y validaciones para facilitar la carga de DJs.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.worksheet.datavalidation import DataValidation
import pandas as pd
from typing import Dict, Any, List, Optional
from pathlib import Path
from ..access_schema import AccessSchema, obtener_metadata


class TemplateGenerator:
    """Generador de plantillas Excel para carga de DJs."""
    
    def __init__(self, metadata: Dict[str, Any], access_schema: AccessSchema = None):
        """
        Inicializa el generador con metadata de la DJ.
        
        Args:
            metadata: Metadata completa de la DJ obtenida de Access
            access_schema: Instancia de AccessSchema para consultas adicionales
        """
        self.metadata = metadata
        self.access_schema = access_schema or AccessSchema()
        self.dj_codigo = metadata['declaracion']['dj_codigo']
        self.nombre_dj = metadata['declaracion']['nombre']
        self.tipo_dj = metadata['declaracion']['tipo']
        self.campos = metadata['campos']
        self.validaciones = metadata['validaciones']
    
    def generar_template(self, output_path: str = None) -> str:
        """
        Genera el template Excel completo.
        
        Args:
            output_path: Ruta donde guardar el archivo. Si no se especifica, usa nombre por defecto.
            
        Returns:
            Ruta del archivo generado
        """
        if output_path is None:
            output_path = f"template_DJ{self.dj_codigo}_{self.nombre_dj.replace(' ', '_')}.xlsx"
        
        wb = Workbook()
        
        if self.tipo_dj == 'SIMPLE':
            self._generar_hoja_simple(wb)
        else:  # COMPUESTA
            self._generar_hojas_compuestas(wb)
        
        # Agregar hoja de instrucciones
        self._agregar_hoja_instrucciones(wb)
        
        # Configurar hoja principal como activa
        if self.tipo_dj == 'SIMPLE':
            wb.active = wb['Datos']
        else:
            # Para compuestas, activar la primera sección
            secciones = self._obtener_secciones()
            if secciones:
                wb.active = wb[secciones[0]]
        
        wb.save(output_path)
        return output_path
    
    def _generar_hoja_simple(self, wb: Workbook) -> None:
        """Genera una sola hoja para DJ simple."""
        # Remover hoja por defecto y crear nueva
        wb.remove(wb.active)
        ws = wb.create_sheet("Datos")
        
        # Obtener campos ordenados por posición
        campos_ordenados = self._obtener_campos_ordenados()
        
        self._configurar_encabezados(ws, campos_ordenados)
        self._agregar_comentarios_y_validaciones(ws, campos_ordenados)
        self._aplicar_estilos(ws, len(campos_ordenados))
    
    def _generar_hojas_compuestas(self, wb: Workbook) -> None:
        """Genera múltiples hojas para DJ compuesta."""
        # Remover hoja por defecto
        wb.remove(wb.active)
        
        # Agrupar campos por sección
        secciones = self._obtener_secciones()
        
        for seccion in secciones:
            ws = wb.create_sheet(seccion)
            campos_seccion = self._obtener_campos_seccion(seccion)
            
            self._configurar_encabezados(ws, campos_seccion)
            self._agregar_comentarios_y_validaciones(ws, campos_seccion)
            self._aplicar_estilos(ws, len(campos_seccion))
    
    def _obtener_campos_ordenados(self) -> List[Dict[str, Any]]:
        """Obtiene lista de campos ordenados por posición."""
        campos_lista = []
        for codigo_campo, info_campo in self.campos.items():
            campos_lista.append({
                'codigo': codigo_campo,
                'info': info_campo
            })
        
        return sorted(campos_lista, key=lambda x: x['info']['posicion'])
    
    def _obtener_secciones(self) -> List[str]:
        """Obtiene lista única de secciones para DJ compuestas."""
        secciones = set()
        for info_campo in self.campos.values():
            if info_campo['seccion']:
                secciones.add(info_campo['seccion'])
        
        return sorted(list(secciones))
    
    def _obtener_campos_seccion(self, seccion: str) -> List[Dict[str, Any]]:
        """Obtiene campos de una sección específica ordenados por posición."""
        campos_seccion = []
        for codigo_campo, info_campo in self.campos.items():
            if info_campo['seccion'] == seccion:
                campos_seccion.append({
                    'codigo': codigo_campo,
                    'info': info_campo
                })
        
        return sorted(campos_seccion, key=lambda x: x['info']['posicion'])
    
    def _configurar_encabezados(self, ws, campos: List[Dict[str, Any]]) -> None:
        """Configura las dos filas de encabezados."""
        # Fila 1: Nombres de campos
        for col, campo in enumerate(campos, 1):
            ws.cell(row=1, column=col, value=campo['info']['nombre'])
        
        # Fila 2: Códigos técnicos (headers del DataFrame)
        for col, campo in enumerate(campos, 1):
            ws.cell(row=2, column=col, value=campo['codigo'])
    
    def _agregar_comentarios_y_validaciones(self, ws, campos: List[Dict[str, Any]]) -> None:
        """Agrega comentarios explicativos y validaciones a las celdas."""
        for col, campo in enumerate(campos, 1):
            codigo_campo = campo['codigo']
            info_campo = campo['info']
            
            # Celda de nombre de campo (fila 1)
            celda_nombre = ws.cell(row=1, column=col)
            
            # Crear comentario con información detallada
            comentario_texto = self._generar_texto_comentario(codigo_campo, info_campo)
            comentario = Comment(comentario_texto, "Sistema DJ")
            comentario.width = 300
            comentario.height = 200
            celda_nombre.comment = comentario
            
            # Agregar validación de datos si aplica
            self._agregar_validacion_columna(ws, col, codigo_campo, info_campo)
    
    def _generar_texto_comentario(self, codigo_campo: str, info_campo: Dict[str, Any]) -> str:
        """Genera el texto del comentario para un campo."""
        comentario_partes = [
            f"CAMPO: {codigo_campo}",
            f"TIPO: {info_campo['tipo_dato']}",
            f"LONGITUD: {info_campo['longitud']}"
        ]
        
        if info_campo['decimales']:
            comentario_partes.append(f"DECIMALES: {info_campo['decimales']}")
        
        if info_campo['obligatorio']:
            comentario_partes.append("⚠️ OBLIGATORIO")
        
        if info_campo['descripcion']:
            comentario_partes.append(f"DESCRIPCIÓN: {info_campo['descripcion']}")
        
        if info_campo['formato_ejemplo']:
            comentario_partes.append(f"EJEMPLO: {info_campo['formato_ejemplo']}")
        
        # Agregar información de validaciones
        if codigo_campo in self.validaciones:
            comentario_partes.append("\nVALIDACIONES:")
            for validacion in self.validaciones[codigo_campo]:
                comentario_partes.append(f"• {validacion['codigo_validacion']}: {validacion['mensaje_error']}")
        
        # Información de tabla lookup
        if info_campo['tabla_lookup']:
            comentario_partes.append(f"\nVALORES VÁLIDOS: Ver tabla {info_campo['tabla_lookup']}")
        
        return "\n".join(comentario_partes)
    
    def _agregar_validacion_columna(self, ws, col: int, codigo_campo: str, info_campo: Dict[str, Any]) -> None:
        """Agrega validación de datos de Excel a una columna."""
        # Validación por longitud máxima (texto)
        if info_campo['tipo_dato'] in ['TEXT', 'VARCHAR', 'CHAR']:
            longitud_max = info_campo['longitud']
            if longitud_max and longitud_max > 0:
                # Crear validación de longitud de texto
                dv = DataValidation(
                    type="textLength",
                    operator="lessThanOrEqual",
                    formula1=str(longitud_max)
                )
                dv.error = f"El texto no puede exceder {longitud_max} caracteres"
                dv.errorTitle = "Longitud inválida"
                dv.prompt = f"Máximo {longitud_max} caracteres"
                dv.promptTitle = f"Campo {codigo_campo}"
                
                # Aplicar a toda la columna (desde fila 3 hasta 1000)
                column_letter = chr(64 + col)  # A=65, B=66, etc.
                dv.add(f"{column_letter}3:{column_letter}1000")
                ws.add_data_validation(dv)
        
        # Validación numérica
        elif info_campo['tipo_dato'] in ['INTEGER', 'DECIMAL', 'NUMERIC']:
            dv = DataValidation(type="decimal")
            dv.error = "Debe ingresar un número válido"
            dv.errorTitle = "Valor inválido"
            dv.prompt = f"Ingrese un número ({info_campo['tipo_dato']})"
            dv.promptTitle = f"Campo {codigo_campo}"
            
            column_letter = chr(64 + col)
            dv.add(f"{column_letter}3:{column_letter}1000")
            ws.add_data_validation(dv)
        
        # Validación de lista para campos con tabla lookup
        if info_campo['tabla_lookup']:
            try:
                valores_lookup = self._obtener_valores_lookup(info_campo['tabla_lookup'])
                if valores_lookup and len(valores_lookup) <= 100:  # Límite de Excel
                    dv = DataValidation(
                        type="list",
                        formula1=f'"{",".join(map(str, valores_lookup))}"'
                    )
                    dv.error = "Seleccione un valor de la lista"
                    dv.errorTitle = "Valor no válido"
                    dv.prompt = "Seleccione de la lista desplegable"
                    dv.promptTitle = f"Campo {codigo_campo}"
                    
                    column_letter = chr(64 + col)
                    dv.add(f"{column_letter}3:{column_letter}1000")
                    ws.add_data_validation(dv)
            except Exception:
                pass  # Si no se puede obtener la lista, continuar sin validación
    
    def _obtener_valores_lookup(self, tabla_lookup: str) -> List[str]:
        """Obtiene valores para validación de lista desde tabla lookup."""
        try:
            df_lookup = self.access_schema.get_tabla_lookup(tabla_lookup)
            # Usar la primera columna como valores
            if len(df_lookup.columns) > 0:
                return df_lookup.iloc[:, 0].astype(str).tolist()
        except Exception:
            pass
        return []
    
    def _aplicar_estilos(self, ws, num_columnas: int) -> None:
        """Aplica estilos visuales a la hoja."""
        # Estilos para encabezados
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Aplicar estilos a encabezados
        for col in range(1, num_columnas + 1):
            # Fila 1: Nombres de campos
            cell1 = ws.cell(row=1, column=col)
            cell1.font = header_font
            cell1.fill = header_fill
            cell1.alignment = header_alignment
            
            # Fila 2: Códigos técnicos
            cell2 = ws.cell(row=2, column=col)
            cell2.font = Font(bold=True, color="000000")
            cell2.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell2.alignment = header_alignment
        
        # Ajustar altura de filas de encabezado
        ws.row_dimensions[1].height = 40
        ws.row_dimensions[2].height = 25
        
        # Ajustar ancho de columnas
        for col in range(1, num_columnas + 1):
            column_letter = chr(64 + col)
            ws.column_dimensions[column_letter].width = 15
        
        # Agregar bordes
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in range(1, 3):  # Solo encabezados
            for col in range(1, num_columnas + 1):
                ws.cell(row=row, column=col).border = thin_border
        
        # Congelar paneles (fijar encabezados)
        ws.freeze_panes = "A3"
    
    def _agregar_hoja_instrucciones(self, wb: Workbook) -> None:
        """Agrega una hoja con instrucciones de uso."""
        ws_inst = wb.create_sheet("Instrucciones", 0)  # Insertar al inicio
        
        # Título
        ws_inst['A1'] = f"INSTRUCCIONES - DJ {self.dj_codigo}: {self.nombre_dj}"
        ws_inst['A1'].font = Font(bold=True, size=16, color="366092")
        
        # Información general
        instrucciones = [
            "",
            "INFORMACIÓN GENERAL:",
            f"• Declaración Jurada: {self.dj_codigo} - {self.nombre_dj}",
            f"• Tipo: {self.tipo_dj}",
            f"• Total de campos: {len(self.campos)}",
            "",
            "INSTRUCCIONES DE USO:",
            "1. Complete los datos a partir de la fila 3",
            "2. Fila 1: Contiene los nombres descriptivos de los campos",
            "3. Fila 2: Contiene los códigos técnicos (NO MODIFIQUE)",
            "4. Pase el cursor sobre los encabezados para ver información detallada",
            "5. Use las validaciones automáticas para evitar errores",
            "",
            "CAMPOS OBLIGATORIOS:",
        ]
        
        # Agregar campos obligatorios
        campos_obligatorios = [
            f"• {codigo}: {info['nombre']}" 
            for codigo, info in self.campos.items() 
            if info['obligatorio']
        ]
        instrucciones.extend(campos_obligatorios)
        
        if not campos_obligatorios:
            instrucciones.append("• No hay campos obligatorios")
        
        instrucciones.extend([
            "",
            "VALIDACIONES ACTIVAS:",
        ])
        
        # Agregar información de validaciones
        total_validaciones = sum(len(v) for v in self.validaciones.values())
        instrucciones.append(f"• Total de validaciones: {total_validaciones}")
        
        for codigo_campo, validaciones_campo in self.validaciones.items():
            nombre_campo = self.campos[codigo_campo]['nombre']
            instrucciones.append(f"• {codigo_campo} ({nombre_campo}): {len(validaciones_campo)} validaciones")
        
        instrucciones.extend([
            "",
            "IMPORTANTE:",
            "• NO modifique la fila 2 (códigos técnicos)",
            "• NO elimine columnas",
            "• Guarde el archivo en formato Excel (.xlsx)",
            "• Use este archivo para cargar datos en el sistema DJ",
            "",
            f"Archivo generado: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ])
        
        # Escribir instrucciones
        for i, linea in enumerate(instrucciones):
            ws_inst[f'A{i+1}'] = linea
            if linea.endswith(":") and linea not in ["", "IMPORTANTE:"]:
                ws_inst[f'A{i+1}'].font = Font(bold=True)
        
        # Ajustar ancho de columna
        ws_inst.column_dimensions['A'].width = 80


def generar_template_dj(dj_codigo: str, output_path: str = None, db_path: str = None) -> str:
    """
    Función de conveniencia para generar template de una DJ.
    
    Args:
        dj_codigo: Código de la declaración jurada
        output_path: Ruta donde guardar el template
        db_path: Ruta opcional al archivo Access
        
    Returns:
        Ruta del archivo generado
    """
    # Obtener metadata
    metadata = obtener_metadata(dj_codigo, db_path)
    
    # Crear generador y generar template
    generator = TemplateGenerator(metadata)
    return generator.generar_template(output_path)


if __name__ == "__main__":
    # Ejemplo de uso
    try:
        print("Generando template para DJ 1922...")
        archivo_generado = generar_template_dj("1922")
        print(f"✓ Template generado: {archivo_generado}")
        
        # Verificar que el archivo existe
        if Path(archivo_generado).exists():
            print(f"✓ Archivo confirmado en: {Path(archivo_generado).absolute()}")
        else:
            print("✗ Error: El archivo no se creó correctamente")
            
    except Exception as e:
        print(f"✗ Error generando template: {e}")
