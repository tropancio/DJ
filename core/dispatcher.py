"""
Módulo dispatcher que orquesta la validación y generación de DJs.
Coordina el flujo completo desde la carga hasta la generación de archivos.
"""

import pandas as pd
from typing import Dict, Any, List, Optional, Union
from pathlib import Path
from datetime import datetime

from .access_schema import AccessSchema, obtener_metadata
from .validation.validator import DJValidator, validar_dj
from .generation.generator_simple import GeneratorSimple, generar_archivo_simple
from .generation.generator_compuesta import GeneratorCompuesta, generar_archivo_compuesto, validar_y_generar_compuesto
from .storage.guardar_access import AccessStorage, guardar_dj_access
from .templates.generar_template import TemplateGenerator, generar_template_dj


class DJDispatcher:
    """Orquestador principal del sistema de DJs."""
    
    def __init__(self, db_path: str = None):
        """
        Inicializa el dispatcher.
        
        Args:
            db_path: Ruta opcional al archivo Access
        """
        self.access_schema = AccessSchema(db_path)
        self.db_path = db_path
    
    def procesar_dj_completo(self, 
                           dj_codigo: str,
                           empresa: Dict[str, Any],
                           datos_entrada: Union[str, pd.DataFrame, Dict[str, pd.DataFrame]],
                           opciones: Dict[str, Any] = None) -> Dict[str, Any]:
        """
        Procesa una DJ completa desde datos de entrada hasta archivo final.
        
        Args:
            dj_codigo: Código de la declaración jurada
            empresa: Datos de la empresa
            datos_entrada: Puede ser:
                - String: ruta a archivo Excel
                - DataFrame: para DJs simples
                - Dict[str, DataFrame]: para DJs compuestas {seccion: dataframe}
            opciones: Opciones de procesamiento
            
        Returns:
            Diccionario con resultado completo del procesamiento
        """
        if opciones is None:
            opciones = {}
        
        resultado = {
            "dj_codigo": dj_codigo,
            "empresa": empresa,
            "inicio_proceso": datetime.now(),
            "pasos_completados": [],
            "errores": [],
            "archivos_generados": {},
            "exito": False
        }
        
        try:
            # 1. Cargar metadata
            print(f"Cargando metadata para DJ {dj_codigo}...")
            metadata = self._cargar_metadata(dj_codigo)
            resultado["metadata"] = metadata
            resultado["pasos_completados"].append("metadata_cargada")
            
            # 2. Procesar datos de entrada
            print("Procesando datos de entrada...")
            datos_procesados = self._procesar_datos_entrada(datos_entrada, metadata)
            resultado["datos_procesados"] = {
                "tipo": type(datos_procesados).__name__,
                "filas_total": self._contar_filas_total(datos_procesados, metadata)
            }
            resultado["pasos_completados"].append("datos_procesados")
            
            # 3. Validar datos
            print("Validando datos...")
            resultado_validacion = self._validar_datos(datos_procesados, metadata)
            resultado["validacion"] = resultado_validacion
            resultado["pasos_completados"].append("datos_validados")
            
            # Generar reporte de errores si hay errores
            if not resultado_validacion["valido"] and opciones.get("generar_reporte_errores", True):
                self._generar_reporte_errores(resultado_validacion, dj_codigo, opciones)
                resultado["pasos_completados"].append("reporte_errores_generado")
            
            # 4. Generar archivo de salida (solo si los datos son válidos o si se fuerza)
            if resultado_validacion["valido"] or opciones.get("forzar_generacion", False):
                print("Generando archivo de salida...")
                archivo_salida = self._generar_archivo_salida(datos_procesados, metadata, empresa, opciones)
                resultado["archivos_generados"]["archivo_sii"] = archivo_salida
                resultado["pasos_completados"].append("archivo_generado")
                
                # 5. Guardar en Access (opcional)
                if opciones.get("guardar_access", False):
                    print("Guardando datos en Access...")
                    resultado_storage = self._guardar_access(datos_procesados, metadata, empresa, opciones)
                    resultado["storage"] = resultado_storage
                    resultado["pasos_completados"].append("datos_guardados")
            
            # Determinar éxito general
            resultado["exito"] = (
                resultado_validacion["valido"] and 
                "archivo_generado" in resultado["pasos_completados"]
            )
            
        except Exception as e:
            resultado["errores"].append(f"Error en procesamiento: {str(e)}")
            resultado["error_critico"] = str(e)
        
        finally:
            resultado["fin_proceso"] = datetime.now()
            resultado["duracion_total"] = (
                resultado["fin_proceso"] - resultado["inicio_proceso"]
            ).total_seconds()
        
        return resultado
    
    def _cargar_metadata(self, dj_codigo: str) -> Dict[str, Any]:
        """Carga metadata de la DJ desde Access."""
        return obtener_metadata(dj_codigo, self.db_path)
    
    def _procesar_datos_entrada(self, datos_entrada: Union[str, pd.DataFrame, Dict[str, pd.DataFrame]], 
                               metadata: Dict[str, Any]) -> Union[pd.DataFrame, Dict[str, pd.DataFrame]]:
        """Procesa los datos de entrada según su tipo."""
        
        if isinstance(datos_entrada, str):
            # Cargar desde archivo Excel
            return self._cargar_desde_excel(datos_entrada, metadata)
        elif isinstance(datos_entrada, pd.DataFrame):
            # DataFrame directo (DJ simple)
            return datos_entrada
        elif isinstance(datos_entrada, dict):
            # Diccionario de DataFrames (DJ compuesta)
            return datos_entrada
        else:
            raise ValueError(f"Tipo de datos de entrada no soportado: {type(datos_entrada)}")
    
    def _cargar_desde_excel(self, ruta_excel: str, metadata: Dict[str, Any]) -> Union[pd.DataFrame, Dict[str, pd.DataFrame]]:
        """Carga datos desde archivo Excel según el tipo de DJ."""
        
        tipo_dj = metadata['declaracion']['tipo']
        
        if tipo_dj == 'SIMPLE':
            # Cargar hoja única
            df = pd.read_excel(ruta_excel, header=1)  # Fila 2 tiene los códigos
            return df
        else:
            # DJ compuesta: cargar múltiples hojas
            dataframes_secciones = {}
            
            # Obtener secciones esperadas
            secciones = set()
            for info_campo in metadata['campos'].values():
                if info_campo['seccion']:
                    secciones.add(info_campo['seccion'])
            
            # Cargar cada sección
            for seccion in secciones:
                try:
                    df_seccion = pd.read_excel(ruta_excel, sheet_name=seccion, header=1)
                    dataframes_secciones[seccion] = df_seccion
                except Exception as e:
                    raise ValueError(f"Error cargando sección '{seccion}': {str(e)}")
            
            return dataframes_secciones
    
    def _contar_filas_total(self, datos: Union[pd.DataFrame, Dict[str, pd.DataFrame]], 
                           metadata: Dict[str, Any]) -> int:
        """Cuenta el total de filas en los datos."""
        if isinstance(datos, pd.DataFrame):
            return len(datos)
        elif isinstance(datos, dict):
            # Para DJ compuesta, usar la primera sección como referencia
            if datos:
                primera_seccion = next(iter(datos.values()))
                return len(primera_seccion)
        return 0
    
    def _validar_datos(self, datos: Union[pd.DataFrame, Dict[str, pd.DataFrame]], 
                      metadata: Dict[str, Any]) -> Dict[str, Any]:
        """Valida los datos según el tipo de DJ."""
        
        tipo_dj = metadata['declaracion']['tipo']
        
        if tipo_dj == 'SIMPLE':
            # Validación directa del DataFrame
            validator = DJValidator(metadata, self.access_schema)
            return validator.validar_dataframe(datos)
        else:
            # DJ compuesta: consolidar y validar
            generator = GeneratorCompuesta(metadata, {})
            
            # Validar estructura de secciones
            validacion_estructura = generator.validar_estructura_secciones(datos)
            
            if validacion_estructura["valido"]:
                # Consolidar y validar contenido
                df_consolidado = generator.consolidar_dataframes(datos)
                validator = DJValidator(metadata, self.access_schema)
                validacion_contenido = validator.validar_dataframe(df_consolidado)
                
                # Combinar resultados
                return {
                    **validacion_contenido,
                    "validacion_estructura": validacion_estructura,
                    "df_consolidado_filas": len(df_consolidado)
                }
            else:
                # Error en estructura, no validar contenido
                return {
                    "valido": False,
                    "errores": validacion_estructura["errores"],
                    "resumen": {"total_filas": 0, "errores_totales": len(validacion_estructura["errores"]), "columnas_con_error": []},
                    "validaciones_omitidas": [],
                    "validacion_estructura": validacion_estructura
                }
    
    def _generar_reporte_errores(self, resultado_validacion: Dict[str, Any], 
                                dj_codigo: str, opciones: Dict[str, Any]) -> str:
        """Genera reporte de errores en Excel."""
        
        output_path = opciones.get("ruta_reporte_errores")
        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = f"errores_DJ{dj_codigo}_{timestamp}.xlsx"
        
        # Usar el validador para generar el reporte
        # (Aquí necesitaríamos instanciar un validador temporal)
        # Por simplicidad, creamos un reporte básico
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Errores"
        
        # Encabezados
        ws['A1'] = 'Fila'
        ws['B1'] = 'Columna'
        ws['C1'] = 'Código Error'
        ws['D1'] = 'Descripción'
        
        # Datos de errores
        for i, error in enumerate(resultado_validacion["errores"], 2):
            ws[f'A{i}'] = error.get("fila", "")
            ws[f'B{i}'] = error.get("columna", "")
            ws[f'C{i}'] = error.get("codigo", "")
            ws[f'D{i}'] = error.get("error", "")
        
        wb.save(output_path)
        return output_path
    
    def _generar_archivo_salida(self, datos: Union[pd.DataFrame, Dict[str, pd.DataFrame]], 
                               metadata: Dict[str, Any], empresa: Dict[str, Any],
                               opciones: Dict[str, Any]) -> str:
        """Genera el archivo de salida según el tipo de DJ."""
        
        tipo_dj = metadata['declaracion']['tipo']
        dj_codigo = metadata['declaracion']['dj_codigo']
        output_path = opciones.get("ruta_archivo_salida")
        
        if tipo_dj == 'SIMPLE':
            return generar_archivo_simple(datos, dj_codigo, empresa, output_path, self.db_path)
        else:
            return generar_archivo_compuesto(datos, dj_codigo, empresa, output_path, self.db_path)
    
    def _guardar_access(self, datos: Union[pd.DataFrame, Dict[str, pd.DataFrame]], 
                       metadata: Dict[str, Any], empresa: Dict[str, Any],
                       opciones: Dict[str, Any]) -> Dict[str, Any]:
        """Guarda los datos en Access."""
        
        tipo_dj = metadata['declaracion']['tipo']
        dj_codigo = metadata['declaracion']['dj_codigo']
        tabla_destino = opciones.get("tabla_destino")
        
        if tipo_dj == 'SIMPLE':
            df_a_guardar = datos
        else:
            # Consolidar DJ compuesta
            generator = GeneratorCompuesta(metadata, empresa)
            df_a_guardar = generator.consolidar_dataframes(datos)
        
        return guardar_dj_access(df_a_guardar, dj_codigo, empresa, tabla_destino, self.db_path)
    
    def generar_template(self, dj_codigo: str, output_path: str = None) -> str:
        """
        Genera template Excel para una DJ.
        
        Args:
            dj_codigo: Código de la DJ
            output_path: Ruta de salida opcional
            
        Returns:
            Ruta del template generado
        """
        return generar_template_dj(dj_codigo, output_path, self.db_path)
    
    def obtener_info_dj(self, dj_codigo: str) -> Dict[str, Any]:
        """
        Obtiene información completa de una DJ.
        
        Args:
            dj_codigo: Código de la DJ
            
        Returns:
            Información estructurada de la DJ
        """
        metadata = self._cargar_metadata(dj_codigo)
        
        info = {
            "declaracion": metadata['declaracion'],
            "resumen": {
                "total_campos": len(metadata['campos']),
                "campos_obligatorios": sum(1 for campo in metadata['campos'].values() if campo['obligatorio']),
                "total_validaciones": sum(len(v) for v in metadata['validaciones'].values()),
                "secciones": []
            },
            "campos_por_seccion": {},
            "validaciones_por_campo": {}
        }
        
        # Agrupar campos por sección
        if metadata['declaracion']['tipo'] == 'COMPUESTA':
            secciones = set()
            for codigo_campo, info_campo in metadata['campos'].items():
                seccion = info_campo['seccion']
                if seccion:
                    secciones.add(seccion)
                    if seccion not in info["campos_por_seccion"]:
                        info["campos_por_seccion"][seccion] = []
                    info["campos_por_seccion"][seccion].append({
                        "codigo": codigo_campo,
                        "nombre": info_campo['nombre'],
                        "obligatorio": info_campo['obligatorio'],
                        "tipo": info_campo['tipo_dato']
                    })
            
            info["resumen"]["secciones"] = sorted(list(secciones))
        
        # Resumen de validaciones por campo
        for codigo_campo, validaciones in metadata['validaciones'].items():
            if validaciones:
                info["validaciones_por_campo"][codigo_campo] = [
                    {
                        "codigo": v["codigo_validacion"],
                        "tipo": v["tipo_validacion"],
                        "mensaje": v["mensaje_error"]
                    }
                    for v in validaciones
                ]
        
        return info


# Funciones de conveniencia
def procesar_dj_desde_excel(ruta_excel: str, dj_codigo: str, empresa: Dict[str, Any],
                           opciones: Dict[str, Any] = None, db_path: str = None) -> Dict[str, Any]:
    """
    Función de conveniencia para procesar DJ desde archivo Excel.
    
    Args:
        ruta_excel: Ruta al archivo Excel
        dj_codigo: Código de la DJ
        empresa: Datos de la empresa
        opciones: Opciones de procesamiento
        db_path: Ruta opcional al archivo Access
        
    Returns:
        Resultado del procesamiento
    """
    dispatcher = DJDispatcher(db_path)
    return dispatcher.procesar_dj_completo(dj_codigo, empresa, ruta_excel, opciones)


def procesar_dj_desde_dataframe(df: pd.DataFrame, dj_codigo: str, empresa: Dict[str, Any],
                               opciones: Dict[str, Any] = None, db_path: str = None) -> Dict[str, Any]:
    """
    Función de conveniencia para procesar DJ desde DataFrame.
    
    Args:
        df: DataFrame con los datos
        dj_codigo: Código de la DJ
        empresa: Datos de la empresa
        opciones: Opciones de procesamiento
        db_path: Ruta opcional al archivo Access
        
    Returns:
        Resultado del procesamiento
    """
    dispatcher = DJDispatcher(db_path)
    return dispatcher.procesar_dj_completo(dj_codigo, empresa, df, opciones)


if __name__ == "__main__":
    # Ejemplo de uso
    import pandas as pd
    
    # Datos de ejemplo
    df_ejemplo = pd.DataFrame({
        'C1': ['12345678-9', '98765432-1'],
        'C2': [1000000, 2500000],
        'C3': ['EMPRESA A', 'EMPRESA B']
    })
    
    # Datos de empresa
    empresa_ejemplo = {
        'rut': '76123456-7',
        'nombre': 'EMPRESA DE PRUEBA S.A.',
        'usuario': 'admin'
    }
    
    # Opciones de procesamiento
    opciones = {
        'generar_reporte_errores': True,
        'guardar_access': True,
        'forzar_generacion': False
    }
    
    try:
        print("Procesando DJ completa...")
        resultado = procesar_dj_desde_dataframe(df_ejemplo, "1922", empresa_ejemplo, opciones)
        
        if resultado["exito"]:
            print("✓ Procesamiento exitoso")
            print(f"  - Pasos completados: {', '.join(resultado['pasos_completados'])}")
            if "archivo_sii" in resultado["archivos_generados"]:
                print(f"  - Archivo generado: {resultado['archivos_generados']['archivo_sii']}")
        else:
            print("✗ Procesamiento con errores:")
            for error in resultado["errores"]:
                print(f"  - {error}")
                
    except Exception as e:
        print(f"✗ Error: {e}")
